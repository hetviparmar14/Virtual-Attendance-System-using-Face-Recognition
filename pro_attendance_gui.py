import tkinter as tk
from tkinter import messagebox, ttk
from deepface import DeepFace
import cv2
import sqlite3
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os
import subprocess
from PIL import Image, ImageTk
import qrcode

def generate_qr(student_name):
    qr_folder = "qr_codes"
    os.makedirs(qr_folder, exist_ok=True)

    qr = qrcode.make(student_name)
    qr.save(f"{qr_folder}/{student_name}.png")
# ---------------- DATABASE ----------------
conn = sqlite3.connect("attendance.db")
cursor = conn.cursor()

cursor.execute("""
CREATE TABLE IF NOT EXISTS attendance(
    name TEXT,
    date TEXT,
    time TEXT,
    lecture TEXT
)
""")
conn.commit()

# ---------------- ADMIN LOGIN ----------------
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "1234"

# ---------------- CAMERA ----------------
cap = None
dataset_path = "dataset"
last_marked = {}

# Face detector
face_cascade = cv2.CascadeClassifier(
    cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
)

# ---------------- CAMERA FUNCTIONS ----------------

def start_camera():
    global cap
    cap = cv2.VideoCapture(0)
    show_frame()

def stop_camera():
    global cap
    if cap:
        cap.release()
        cv2.destroyAllWindows()

def show_frame():
    global cap, last_marked

    if cap is None:
        return

    ret, frame = cap.read()
    if not ret:
        return

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    faces = face_cascade.detectMultiScale(
        gray,
        scaleFactor=1.2,
        minNeighbors=5
    )

    for (x, y, w, h) in faces:

        face_img = frame[y:y+h, x:x+w]
        cv2.imwrite("temp_face.jpg", face_img)

        name = "Unknown"

        try:
            result = DeepFace.find(
                img_path="temp_face.jpg",
                db_path=dataset_path,
                enforce_detection=False
            )

            if len(result[0]) > 0:

                distance = result[0]['distance'][0]
               


                # 🔥 STRICT THRESHOLD (adjust if needed)
                if distance < 0.60:
                    path = result[0]['identity'][0]
                    name = os.path.basename(os.path.dirname(path))
                    show_student_preview(name, "temp_face.jpg")

                else:
                    name = "Unknown"

                # Save only if recognized
                if name != "Unknown":

                    now = datetime.now()
                    today = now.strftime("%d-%m-%Y")
                    time_now = now.strftime("%H:%M:%S")
                    lecture = lecture_entry.get()

                    if name not in last_marked:
                        last_marked[name] = True

                        cursor.execute(
                            "INSERT INTO attendance VALUES (?,?,?,?)",
                            (name, today, time_now, lecture)
                        )
                        conn.commit()

                        save_to_excel(name, today, time_now, lecture)
                        update_report()

        except Exception as e:
            print("Face error:", e)

        # Draw box
        color = (0, 255, 0) if name != "Unknown" else (0, 0, 255)

        cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)
        cv2.putText(frame, name, (x, y-10),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.8, color, 2)

    cv2.imshow("Virtual Attendance", frame)
    root.after(30, show_frame)

# ---------------- EXCEL ----------------

def save_to_excel(name, date, time_now, lecture):
    excel_file = "attendance.xlsx"

    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Date", "Time", "Lecture"])
        wb.save(excel_file)

    wb = load_workbook(excel_file)
    ws = wb.active
    ws.append([name, date, time_now, lecture])
    wb.save(excel_file)

def open_excel():
    excel_file = "attendance.xlsx"
    if os.path.exists(excel_file):
        subprocess.Popen(excel_file, shell=True)
    else:
        messagebox.showerror("Error", "Excel file not found")

# ---------------- REPORT ----------------

def update_report():
    today = datetime.now().strftime("%d-%m-%Y")

    cursor.execute(
        "SELECT COUNT(DISTINCT name) FROM attendance WHERE date=?",
        (today,)
    )
    total = cursor.fetchone()[0]

    total_label.config(text=f"Total Students Today: {total}")
    present_label.config(text=f"Present Count: {total}")

# ---------------- ADMIN ----------------

def admin_login():
    login_win = tk.Toplevel(root)
    login_win.title("Admin Login")
    login_win.geometry("300x200")

    tk.Label(login_win, text="Username").pack(pady=5)
    username_entry = tk.Entry(login_win)
    username_entry.pack()

    tk.Label(login_win, text="Password").pack(pady=5)
    password_entry = tk.Entry(login_win, show="*")
    password_entry.pack()

    def check_login():
        if (username_entry.get() == ADMIN_USERNAME and
            password_entry.get() == ADMIN_PASSWORD):
            login_win.destroy()
            open_admin_panel()
        else:
            messagebox.showerror("Error", "Invalid Credentials")

    tk.Button(login_win, text="Login", command=check_login).pack(pady=15)

def open_admin_panel():
    admin_win = tk.Toplevel(root)
    admin_win.title("Admin Dashboard")
    admin_win.geometry("350x350")

    tk.Label(admin_win,
             text="Admin Dashboard",
             font=("Arial",14,"bold")).pack(pady=15)

    tk.Button(admin_win, text="View Attendance",
              command=view_attendance,
              width=25).pack(pady=5)

    tk.Button(admin_win, text="Open Excel File",
              command=open_excel,
              width=25).pack(pady=5)

    tk.Button(admin_win, text="Attendance Tracker",
              command=open_tracker,
              width=25).pack(pady=5)

    tk.Button(admin_win, text="Clear Today's Attendance",
              command=clear_today_attendance,
              width=25,
              bg="red", fg="white").pack(pady=10)

def clear_today_attendance():
    today = datetime.now().strftime("%d-%m-%Y")
    cursor.execute("DELETE FROM attendance WHERE date=?", (today,))
    conn.commit()
    messagebox.showinfo("Success", "Today's attendance cleared")
    update_report()

# ---------------- VIEW ATTENDANCE ----------------

def view_attendance():
    view_win = tk.Toplevel(root)
    view_win.title("Attendance Records")
    view_win.geometry("500x300")

    tree = ttk.Treeview(
        view_win,
        columns=("Name","Date","Time","Lecture"),
        show="headings"
    )

    for col in ("Name","Date","Time","Lecture"):
        tree.heading(col, text=col)

    tree.pack(fill="both", expand=True)

    cursor.execute("SELECT * FROM attendance")
    rows = cursor.fetchall()

    for row in rows:
        tree.insert("", tk.END, values=row)

# ---------------- ATTENDANCE TRACKER ----------------

def open_tracker():
    tracker_win = tk.Toplevel(root)
    tracker_win.title("Attendance Tracker")
    tracker_win.geometry("400x300")

    today = datetime.now().strftime("%d-%m-%Y")

    cursor.execute("SELECT COUNT(DISTINCT name) FROM attendance")
    total_students = cursor.fetchone()[0]

    cursor.execute(
        "SELECT COUNT(DISTINCT name) FROM attendance WHERE date=?",
        (today,)
    )
    present_today = cursor.fetchone()[0]

    absent_today = total_students - present_today
    percentage = (present_today / total_students * 100) if total_students > 0 else 0

    tk.Label(tracker_win,
             text="Attendance Tracker",
             font=("Arial",14,"bold")).pack(pady=15)

    tk.Label(tracker_win, text=f"Total Students: {total_students}").pack(pady=5)
    tk.Label(tracker_win, text=f"Present Today: {present_today}").pack(pady=5)
    tk.Label(tracker_win, text=f"Absent Today: {absent_today}").pack(pady=5)
    tk.Label(tracker_win,
             text=f"Attendance Percentage: {percentage:.2f}%",
             font=("Arial",11,"bold")).pack(pady=10)

# ---------------- STUDENT REGISTRATION ----------------

def register_student():

    reg_win = tk.Toplevel(root)
    reg_win.title("Register Student")
    reg_win.geometry("300x200")

    tk.Label(reg_win, text="Enter Student Name",
             font=("Arial",12)).pack(pady=10)

    name_entry = tk.Entry(reg_win, font=("Arial",12))
    name_entry.pack(pady=5)

    def start_capture():

        student_name = name_entry.get().strip()

        if student_name == "":
            messagebox.showerror("Error", "Enter student name")
            return

        # Create folder inside dataset
        student_path = os.path.join(dataset_path, student_name)

        if not os.path.exists(student_path):
            os.makedirs(student_path)

        cap = cv2.VideoCapture(0)
        count = 0

        messagebox.showinfo("Info", "Capturing 15 images... Look at camera")

        while count < 15:
            ret, frame = cap.read()
            if not ret:
                break

            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray, 1.3, 5)

            for (x, y, w, h) in faces:
                face = frame[y:y+h, x:x+w]
                file_path = os.path.join(student_path, f"{count}.jpg")
                cv2.imwrite(file_path, face)
                count += 1

                cv2.rectangle(frame, (x,y), (x+w,y+h), (0,255,0), 2)

            cv2.imshow("Registering Face", frame)

            if cv2.waitKey(1) == 13:
                break

        cap.release()
        cv2.destroyAllWindows()

        messagebox.showinfo("Success", "Student Registered Successfully")
        reg_win.destroy()

    tk.Button(reg_win, text="Start Capture",
              command=start_capture,
              width=20).pack(pady=15)
def show_student_preview(name, face_img_path):
    try:
        img = Image.open(face_img_path)
        img = img.resize((120, 120))
        img = ImageTk.PhotoImage(img)

        photo_label.config(image=img)
        photo_label.image = img

        student_name_label.config(text=f"Detected: {name}")

    except Exception as e:
        print("Preview error:", e)

# ---------------- MAIN GUI ----------------

root = tk.Tk()
root.title("Virtual Attendance System")
root.geometry("400x520")
root.resizable(False, False)

# ---------------- MODERN STYLE ----------------

style = ttk.Style()
style.theme_use("clam")

# Colors
PRIMARY = "#1f2937"      # Main background
SECONDARY = "#111827"    # Card background
ACCENT = "#3b82f6"       # Blue
SUCCESS = "#10b981"
DANGER = "#ef4444"
TEXT = "#ffffff"

root.configure(bg=PRIMARY)

# Button styling
style.configure("TButton",
                font=("Segoe UI", 11),
                padding=8)

style.configure("Accent.TButton",
                background=ACCENT,
                foreground="white")

style.map("Accent.TButton",
          background=[("active", "#2563eb")])

# ---------------- TITLE ----------------

tk.Label(root,
         text="🎓 Virtual Attendance System",
         font=("Segoe UI", 16, "bold"),
         bg=PRIMARY,
         fg=TEXT).pack(pady=20)

# ---------------- LECTURE ENTRY ----------------

tk.Label(root,
         text="Lecture Name",
         font=("Segoe UI", 11),
         bg=PRIMARY,
         fg=TEXT).pack(pady=5)

lecture_entry = tk.Entry(root,
                         font=("Segoe UI", 11),
                         width=25,
                         bd=0,
                         relief="flat")
lecture_entry.pack(pady=10, ipady=6)

# ---------------- BUTTONS FRAME ----------------

button_frame = tk.Frame(root, bg=PRIMARY)
button_frame.pack(pady=10)

ttk.Button(button_frame,
           text="Start Camera",
           command=start_camera,
           style="Accent.TButton",
           width=22).pack(pady=6)

ttk.Button(button_frame,
           text="Stop Camera",
           command=stop_camera,
           width=22).pack(pady=6)

ttk.Button(button_frame,
           text="Register Student",
           command=register_student,
           width=22).pack(pady=6)

ttk.Button(button_frame,
           text="Admin Login",
           command=admin_login,
           width=22).pack(pady=6)

# ---------------- STATS CARD ----------------

stats_frame = tk.Frame(root,
                       bg=SECONDARY,
                       bd=0)
stats_frame.pack(pady=25, fill="x", padx=30)

tk.Label(stats_frame,
         text="📊 Today's Statistics",
         font=("Segoe UI", 12, "bold"),
         bg=SECONDARY,
         fg=TEXT).pack(pady=10)

total_label = tk.Label(stats_frame,
                       text="Total Students Today: 0",
                       font=("Segoe UI", 10),
                       bg=SECONDARY,
                       fg=TEXT)
total_label.pack(pady=5)
# Student Photo Preview
photo_label = tk.Label(stats_frame,
                       bg=SECONDARY)
photo_label.pack(pady=10)

student_name_label = tk.Label(stats_frame,
                              text="No Student Detected",
                              font=("Segoe UI", 10, "bold"),
                              bg=SECONDARY,
                              fg="#facc15")
student_name_label.pack(pady=5)

present_label = tk.Label(stats_frame,
                         text="Present Count: 0",
                         font=("Segoe UI", 10),
                         bg=SECONDARY,
                         fg=TEXT)
present_label.pack(pady=5)

# ---------------- FOOTER ----------------

tk.Label(root,
         text="AI Powered Face Recognition",
         font=("Segoe UI", 9),
         bg=PRIMARY,
         fg="#9ca3af").pack(side="bottom", pady=10)

root.mainloop()
